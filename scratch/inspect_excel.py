import openpyxl

wb = openpyxl.load_workbook("datos/workexample.xlsx")
print("Sheets:", wb.sheetnames)
ws = wb.active

for r in range(4, 20):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 32)]
    if any(row_vals):
        print(f"Row {r}: {row_vals}")
