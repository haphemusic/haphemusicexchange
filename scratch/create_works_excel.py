import openpyxl

# Create a brand new workbook instead of loading the human-friendly template.
# This ensures SheetJS parses it with Row 1 as headers correctly.
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Obras"

# Headers exactly matching the column names expected by the parser
headers = [
    "Original Title",
    "Sub-title / Version",
    "Year of Composition",
    "Duration",
    "Catalogue Number",
    "Scoring Category",
    "Number of Performers",
    "Detailed Instrumentation",
    "Soloist Instrument",
    "Unusual Objects / Preparations",
    "Use of Electronics",
    "Type of Electronics",
    "Required Software / Tech",
    "Premiere Date",
    "Premiere Venue",
    "Premiere City",
    "Premiere Performers",
    "Commissioning",
    "Publisher",
    "Score Status",
    "Availability of Materials",
    "Copyright Society",
    "Score Sample (Link/PDF)",
    "Audio/Video Link",
    "Recording Type",
    "Program Notes",
    "Aesthetic / Style Tags",
    "Language / Librettist",
    "Technical Difficulty",
    "Space Requirements",
    "Info+"
]

# Write headers in Row 1
for col_idx, h in enumerate(headers):
    ws.cell(row=1, column=col_idx + 1, value=h)

# Define the 5 works data rows (31 columns each)
works = [
    # 1. Ligeti - Atmosphères
    [
        "Atmosphères",                         # Original Title
        "",                                    # Sub-title / Version
        1961,                                  # Year of Composition
        9.0,                                   # Duration
        "",                                    # Catalogue Number
        "orchestra",                           # Scoring Category
        87,                                    # Number of Performers
        "4 Fl, 4 Ob, 4 Cl, 3 Bsn, 6 Hn, 4 Tpt, 4 Tbn, 1 Tba, Pno, Vln I (14), Vln II (14), Vla (10), Vc (10), Cb (8)", # Detailed Instrumentation
        "",                                    # Soloist Instrument
        "",                                    # Unusual Objects / Preparations
        "No",                                  # Use of Electronics
        "",                                    # Type of Electronics
        "",                                    # Required Software / Tech
        "22/10/1961",                          # Premiere Date
        "Donaueschingen Festival",             # Premiere Venue
        "Donaueschingen",                      # Premiere City
        "Südwestfunk-Orchester, Hans Rosbaud", # Premiere Performers
        "SWF",                                 # Commissioning
        "Peters",                              # Publisher
        "finished",                            # Score Status
        "rental_only",                         # Availability of Materials
        "GEMA",                                # Copyright Society
        "https://www.edition-peters.com/product/atmospheres/ep5941", # Score Sample
        "https://www.youtube.com/watch?v=aI0P1NnUFxc",               # Audio/Video Link
        "studio",                              # Recording Type
        "A landmark of sound-mass composition, focusing on texture and timbre rather than melody or rhythm.", # Program Notes
        "micropolyphony, sound-mass, spectralism-precursor", # Aesthetic / Style Tags
        "",                                    # Language / Librettist
        "professional",                        # Technical Difficulty
        "Large orchestral stage",              # Space Requirements
        ""                                     # Info+
    ],
    # 2. Varèse - Déserts
    [
        "Déserts",
        "For wind, percussion and electronic tape",
        1954,
        24.0,
        "",
        "ensemble",
        20,
        "2 Fl, Ob, 2 Cl, 3 Hn, 3 Tpt, 3 Tbn, Tba, Perc (5), Tape",
        "",
        "Tape playback",
        "Yes",
        "fixed_media",
        "2-channel tape playback system",
        "02/12/1954",
        "Théâtre des Champs-Élysées",
        "Paris",
        "Orchestre National de France, Hermann Scherchen",
        "",
        "Colfranc",
        "finished",
        "sale_or_rental",
        "ASCAP",
        "",
        "https://www.youtube.com/watch?v=7MizQf14JpE",
        "concert",
        "One of the earliest major compositions combining live instrumental performance with electronic music on magnetic tape.",
        "electronic, wind-ensemble, tape-music",
        "",
        "professional",
        "Standard stage with stereo PA system",
        ""
    ],
    # 3. Satie - Vexations
    [
        "Vexations",
        "",
        1893,
        1080.0,
        "",
        "solo",
        1,
        "Pno",
        "Piano",
        "",
        "No",
        "",
        "",
        "09/09/1963",
        "Pocket Theatre",
        "New York",
        "John Cage and team of relay pianists",
        "",
        "Max Eschig",
        "finished",
        "public_domain",
        "SACEM",
        "",
        "https://www.youtube.com/watch?v=F0f19mN9t8Y",
        "concert",
        "A single short musical theme repeated 840 times. Designed to test the endurance of both performer and listener.",
        "minimialism, conceptual-art, endurance",
        "",
        "student",
        "Regular piano room",
        "Often performed by multiple pianists in shifts."
    ],
    # 4. Penderecki - Threnody to the Victims of Hiroshima
    [
        "Threnody to the Victims of Hiroshima",
        "",
        1960,
        8.5,
        "",
        "ensemble",
        52,
        "Vln I (24), Vla (10), Vc (10), Cb (8)",
        "",
        "Extended string techniques (bowing behind bridge, striking body of instrument)",
        "No",
        "",
        "",
        "22/09/1960",
        "Warsaw Autumn Festival",
        "Warsaw",
        "Warsaw Philharmonic Orchestra, Jan Krenz",
        "",
        "PWM / Schott",
        "finished",
        "sale_or_rental",
        "ZAIKS",
        "https://www.schott-music.com/en/threnody-to-the-victims-of-hiroshima-noc381710.html",
        "https://www.youtube.com/watch?v=Dp3BlFZWJNA",
        "studio",
        "Written for 52 string instruments, employing graphic notation and extended techniques to create screeching clusters.",
        "sonorism, microtonal, graphic-notation",
        "",
        "professional",
        "Standard stage",
        "Originally titled 8'37\"."
    ],
    # 5. John Chowning - Stria
    [
        "Stria",
        "",
        1977,
        16.5,
        "",
        "other",
        0,
        "Computer generated tape",
        "",
        "",
        "Yes",
        "fixed_media",
        "Quadraphonic sound system",
        "13/10/1977",
        "IRCAM",
        "Paris",
        "Computer playback",
        "IRCAM",
        "",
        "finished",
        "rental_only",
        "BMI",
        "",
        "https://www.youtube.com/watch?v=Fnt2RocG17k",
        "studio",
        "A landmark electronic piece generated using FM synthesis based on the golden ratio.",
        "computer-music, fm-synthesis, algorithmic",
        "",
        "advanced",
        "Hall with quadraphonic PA system",
        ""
    ]
]

# Write data in Row 2 to 6
for i, work_data in enumerate(works):
    current_row = 2 + i
    for col_idx, val in enumerate(work_data):
        ws.cell(row=current_row, column=col_idx + 1, value=val)

output_file = "datos/obras_ejemplo.xlsx"
wb.save(output_file)
print(f"Successfully generated clean, flat {output_file} with {len(works)} works!")
